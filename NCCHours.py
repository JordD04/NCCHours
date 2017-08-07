# Program that tracks hours

import openpyxl, shelve, datetime
from openpyxl.utils import get_column_letter

monthsD = {'may': 31, 'june': 30, 'july': 31, 'august': 31, 'september': 30, 'october': 31, 'november': 30, 'december': 31}                 # creates a dictionary of months
monthsL = ['january', 'february', 'march', 'april', 'may', 'june', 'july', 'august', 'september', 'october', 'november', 'december']        # creates a list of months

while 1 == 1:                                                           # loop

    shelfFile = shelve.open('Variable', writeback=True)                 # sets a shelf variable to determine if time is being tracked
    var = shelfFile['variable']                                         # defines var as shelf variable
    
    command = input('Enter: start, or stop: ')                          # command input

    

    if command == 'start':
        shelfFile = shelve.open('Variable')
        if var == 0:
            print('Time Already Running')
        else: 
            dt = datetime.datetime.now()                                # finds the datetime
            month = dt.month                                            # finds the month
            day = dt.day                                                # finds the day
            time = dt.time()                                            # finds the time
            weekday = dt.weekday()                                      # finds the week day
        
            wb = openpyxl.load_workbook('hours.xlsx')                   # opens spreadsheet
            
                                                                        # determines the current week number
            month2 = month                                              # creates a new variable to track addition of days of all months
            totalDays = 1                                               # variable to track total number of days in previous months
            while month2 > 5:                                           # loops until the total number of all days for each passed month has been added
                month2 = month2-1                                       # sets the month2 value to the oldest uncounted month   
                totalDays = totalDays + monthsD[monthsL[month2]]        # adds the total number of days in the oldest uncounted month to the total number of days
            totalDays = totalDays + day                                 # adds the number of days passed in this month to the total number of passed days
            week = int(totalDays / 7)
            
            sheetName = ''.join(['week',str(week)])                     # creates name of sheet by concatenating 'week' and the week number           

            sheet = wb.get_sheet_by_name(sheetName)                     # references sheet of this week

            cellDay = weekday+(3*(weekday))+1                           # general formula to determine input column based on day
            cellRowNoCell = ''.join([get_column_letter(cellDay+3),'1']) # determines coordinates of cell containing the number of rows used for that day, cellRowNo is used to indicate the next row which is empty                                                 
            cellRowNo = sheet[cellRowNoCell].value                      # reads value of cell used to determine the number of rows for that day
            if cellRowNo:                                               # sets the number of cells to 0 if it's a new day
                print('Back to Work!')
            else:
                sheet[cellRowNoCell] = 0
                cellRowNo = sheet[cellRowNoCell].value

            cellLetter = get_column_letter(cellDay+2)                   # letter of cell
            cellNumber = cellRowNo+4                                    # number of cell
            cell = ''.join([cellLetter, str(cellNumber)])               # determines the position of the cell for the start time
            sheet[cell] = time

            print(sheet)
            print('Cell: ' + cell)
            print('Time: ' + str(time))
         
            shelfFile['variable'] = 0                                   # sets variable to indicate that counting is in progress
            wb.save('hours.xlsx')                                       # saves the spreadsheet



    if command == 'stop':
        shelfFile = shelve.open('Variable')

        if var == 1:
            print('Nothing to stop!')
        else:
            dt = datetime.datetime.now()                                # finds the datetime
            month = dt.month                                            # finds the month
            day = dt.day                                                # finds the day
            time = dt.time()                                            # finds the time
            weekday = dt.weekday()                                      # finds the week day
        
            wb = openpyxl.load_workbook('hours.xlsx')                   # opens spreadsheet

                                                                        # determines the current week number
            month2 = month                                              # creates a new variable to track addition of days of all months
            totalDays = 1                                               # variable to track total number of days in previous months
            while month2 > 5:                                           # loops until the total number of all days for each passed month has been added. Counting started in may, hence 5
                month2 = month2-1                                       # sets the month2 value to the oldest uncounted month   
                totalDays = totalDays + monthsD[monthsL[month2]]        # adds the total number of days in the oldest uncounted month to the total number of days
            totalDays = totalDays + day                                 # adds the number of days passed in this month to the total number of passed days
            week = int(totalDays / 7)
            
            sheetName = ''.join(['week',str(week)])                     # creates name of sheet by concatenating 'week' and the week number           

            sheet = wb.get_sheet_by_name(sheetName)                     # references sheet of this week

            cellDay = weekday+(3*(weekday))+1                           # general formula to determine input column based on day
            cellRowNoCell = ''.join([get_column_letter(cellDay+3),'1']) # determines coordinates of cell containing the number of rows used for that day, cellRowNo is used to indicate the next row which is empty                                                 
            cellRowNo = sheet[cellRowNoCell].value                      # reads value of cell used to determine the number of rows for that day

            cellLetter = get_column_letter(cellDay+3)                   # letter of cell
            cellNumber = cellRowNo+4                                    # number of cell
            cell = ''.join([cellLetter, str(cellNumber)])               # determines the position of the cell for the start time
            sheet[cell] = time

            sheet[cellRowNoCell] = cellRowNo+1                          # increments cellRowNo values by 1

            print(sheet)
            print('Cell: ' + cell)
            print('Time: ' + str(time))

            shelfFile['variable'] = 1                                   # sets variable to indicate that counting has stopped
            wb.save('hours.xlsx')                                       # saves the spreadsheet
